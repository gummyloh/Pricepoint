from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import io
import logging
import uuid
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Annotated, Any
from contextlib import asynccontextmanager

import bcrypt
import jwt
import pandas as pd
import asyncpg
from fastapi import (
    FastAPI,
    APIRouter,
    HTTPException,
    Depends,
    Request,
    Response,
    UploadFile,
    File,
    Query,
)
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import StreamingResponse
from pydantic import BaseModel, Field, EmailStr, BeforeValidator, ConfigDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from schneider_pdf import parse_schneider_pdf

# ---------- Postgres pool ----------
DATABASE_URL = os.environ["DATABASE_URL"]

_pool: Optional[asyncpg.Pool] = None


async def get_pool() -> asyncpg.Pool:
    global _pool
    if _pool is None:
        _pool = await asyncpg.create_pool(
            DATABASE_URL,
            min_size=1,
            max_size=10,
            # Supabase pooler uses PgBouncer transaction mode — disable prepared statement caching.
            statement_cache_size=0,
        )
    return _pool


# ---------- Constants ----------
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_TTL_MINUTES = 60 * 12
REFRESH_TOKEN_TTL_DAYS = 7


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def _cookie_secure() -> bool:
    return os.environ.get("COOKIE_SECURE", "false").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _cookie_samesite() -> str:
    v = os.environ.get("COOKIE_SAMESITE", "lax").strip().lower()
    return v if v in ("lax", "strict", "none") else "lax"


# ---------- Models ----------
def _uuid_to_str(v: Any) -> str:
    if isinstance(v, uuid.UUID):
        return str(v)
    return str(v)


PyUUID = Annotated[str, BeforeValidator(_uuid_to_str)]


class UserPublic(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    id: PyUUID = Field(alias="_id")
    email: EmailStr
    name: str
    role: str
    principal: Optional[str] = None
    created_at: Optional[datetime] = None


class RegisterInput(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginInput(BaseModel):
    email: EmailStr
    password: str


class InviteUserInput(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)
    role: str = Field(default="admin")
    principal: Optional[str] = None  # required when role="admin"; ignored for role="master_admin"


class PriceRecordCreate(BaseModel):
    part_no: str = Field(min_length=1)
    unit_price: float = Field(ge=0)
    cpq_number: str = Field(min_length=1)
    cpq_date: str
    customer: str = Field(min_length=1)
    cpq_price: float = Field(ge=0)
    qty: int = Field(default=1, ge=1)
    description: Optional[str] = ""
    notes: Optional[str] = ""
    principal: Optional[str] = None  # master_admin only; ignored for principal-scoped admins


class PriceRecordUpdate(BaseModel):
    part_no: Optional[str] = None
    unit_price: Optional[float] = None
    cpq_number: Optional[str] = None
    cpq_date: Optional[str] = None
    customer: Optional[str] = None
    cpq_price: Optional[float] = None
    qty: Optional[int] = None
    description: Optional[str] = None
    notes: Optional[str] = None


class CPQBatchLine(BaseModel):
    part_no: str
    unit_price: float
    customer: str
    cpq_price: float
    qty: int = Field(default=1, ge=1)
    description: Optional[str] = ""
    notes: Optional[str] = ""


class CPQBatchCreate(BaseModel):
    cpq_number: str
    cpq_date: str
    lines: List[CPQBatchLine]
    principal: Optional[str] = None  # master_admin only; ignored for principal-scoped admins


class ImportRow(BaseModel):
    part_no: str
    unit_price: Optional[str] = None
    cpq_number: str
    cpq_date: str
    customer: str
    cpq_price: Optional[str] = None
    qty: Optional[str] = None
    description: Optional[str] = ""
    notes: Optional[str] = ""


class ImportCommit(BaseModel):
    rows: List[ImportRow]
    principal: Optional[str] = None  # master_admin only; ignored for principal-scoped admins


class DuplicateCPQInput(BaseModel):
    cpq_number: str
    source_customer: str
    target_customers: List[str] = Field(min_length=1)
    new_cpq_number: Optional[str] = None
    new_cpq_date: Optional[str] = None


# ---------- Password Helpers ----------
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


# ---------- JWT Helpers ----------
def create_access_token(user_id: str, email: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_TTL_MINUTES),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_TTL_DAYS),
        "type": "refresh",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def set_auth_cookies(response: Response, access: str, refresh: str) -> None:
    secure = _cookie_secure()
    samesite = _cookie_samesite()
    response.set_cookie(
        key="access_token",
        value=access,
        httponly=True,
        secure=secure,
        samesite=samesite,
        max_age=ACCESS_TOKEN_TTL_MINUTES * 60,
        path="/",
    )
    response.set_cookie(
        key="refresh_token",
        value=refresh,
        httponly=True,
        secure=secure,
        samesite=samesite,
        max_age=REFRESH_TOKEN_TTL_DAYS * 24 * 3600,
        path="/",
    )


def clear_auth_cookies(response: Response) -> None:
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT id, email, name, role, principal, created_at FROM users WHERE id = $1",
        uuid.UUID(user_id),
    )
    if not row:
        raise HTTPException(status_code=401, detail="User not found")
    return {
        "_id": str(row["id"]),
        "email": row["email"],
        "name": row["name"],
        "role": row["role"],
        "principal": row["principal"],
        "created_at": row["created_at"],
    }


# ---------- Serialization Helpers ----------
def compute_discount(unit_price: float, cpq_price: float) -> float:
    if unit_price and unit_price > 0:
        return round((unit_price - cpq_price) / unit_price * 100, 2)
    return 0.0


def _fmt_date(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v)
    return s[:10]


def serialize_price_record(row: Any) -> dict:
    """Accepts an asyncpg Record or dict-like."""
    unit_price = float(row["unit_price"]) if row["unit_price"] is not None else 0.0
    cpq_price = float(row["cpq_price"]) if row["cpq_price"] is not None else 0.0
    return {
        "id": str(row["id"]),
        "part_no": row["part_no"] or "",
        "unit_price": unit_price,
        "cpq_number": row["cpq_number"] or "",
        "cpq_date": _fmt_date(row["cpq_date"]),
        "customer": row["customer"] or "",
        "cpq_price": cpq_price,
        "discount_pct": compute_discount(unit_price, cpq_price),
        "qty": int(row["qty"]) if row["qty"] is not None else 1,
        "description": row["description"] or "",
        "notes": row["notes"] or "",
        "principal": row["principal"],
        "created_by": str(row["created_by"]) if row["created_by"] else None,
        "created_by_name": row["created_by_name"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def parse_qty(v: Any) -> int:
    """Parse a possibly-blank raw qty (from import rows) into an int, default 1."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return 1
    return int(float(v))


def parse_iso_date(s: str) -> date:
    """Parse various date formats → python date."""
    try:
        if isinstance(s, date) and not isinstance(s, datetime):
            return s
        if isinstance(s, datetime):
            return s.date()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(str(s)[:10], fmt).date()
            except ValueError:
                continue
        return pd.to_datetime(s).date()
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid date: {s}")


# ---------- Multi-tenancy Helpers ----------
VALID_PRINCIPALS = {"Farg", "Hexmatic", "Nixma"}


def resolve_principal(user: dict, requested: Optional[str] = None) -> str:
    """Derive the principal to stamp on a new price_records row.

    Principal-scoped admins are always tied to their own principal — any
    client-supplied `principal` is ignored for them. master_admin has no
    principal of their own and must supply one explicitly.
    """
    if user["role"] == "master_admin":
        if not requested or requested not in VALID_PRINCIPALS:
            raise HTTPException(
                status_code=400,
                detail=f"principal is required for master_admin and must be one of {sorted(VALID_PRINCIPALS)}",
            )
        return requested
    return user["principal"]


def principal_filter_clause(user: dict, param_idx: int) -> tuple[str, list]:
    """SQL fragment (no leading WHERE/AND) + params scoping a query to the
    user's principal. Returns ("", []) for master_admin, who is unscoped."""
    if user["role"] == "master_admin":
        return "", []
    return f"principal = ${param_idx}", [user["principal"]]


# ---------- App ----------
@asynccontextmanager
async def lifespan(app: FastAPI):
    await get_pool()
    yield
    global _pool
    if _pool is not None:
        await _pool.close()
        _pool = None


app = FastAPI(lifespan=lifespan)
api = APIRouter(prefix="/api")


# ---------- Auth Routes ----------
@api.get("/auth/bootstrap-status")
async def bootstrap_status():
    pool = await get_pool()
    n = await pool.fetchval("SELECT COUNT(*) FROM users")
    return {"has_users": (n or 0) > 0}


@api.post("/auth/register", response_model=UserPublic)
async def register(input: RegisterInput, response: Response):
    pool = await get_pool()
    n = await pool.fetchval("SELECT COUNT(*) FROM users")
    if (n or 0) > 0:
        raise HTTPException(
            status_code=403,
            detail="Registration is closed. Ask an admin to invite you.",
        )
    email = input.email.lower()
    existing = await pool.fetchval("SELECT 1 FROM users WHERE email = $1", email)
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")

    row = await pool.fetchrow(
        """
        INSERT INTO users (email, password_hash, name, role, principal, created_at)
        VALUES ($1, $2, $3, 'master_admin', NULL, now())
        RETURNING id, email, name, role, principal, created_at
        """,
        email,
        hash_password(input.password),
        input.name,
    )
    uid = str(row["id"])
    access = create_access_token(uid, email)
    refresh = create_refresh_token(uid)
    set_auth_cookies(response, access, refresh)
    return {
        "_id": uid,
        "email": row["email"],
        "name": row["name"],
        "role": row["role"],
        "principal": row["principal"],
        "created_at": row["created_at"],
    }


@api.post("/auth/login", response_model=UserPublic)
async def login(input: LoginInput, response: Response):
    email = input.email.lower()
    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT id, email, password_hash, name, role, principal, created_at FROM users WHERE email = $1",
        email,
    )
    if not row or not verify_password(input.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    uid = str(row["id"])
    access = create_access_token(uid, email)
    refresh = create_refresh_token(uid)
    set_auth_cookies(response, access, refresh)
    return {
        "_id": uid,
        "email": row["email"],
        "name": row["name"],
        "role": row["role"],
        "principal": row["principal"],
        "created_at": row["created_at"],
    }


@api.post("/auth/logout")
async def logout(response: Response):
    clear_auth_cookies(response)
    return {"ok": True}


@api.get("/auth/me", response_model=UserPublic)
async def me(user: dict = Depends(get_current_user)):
    return user


@api.post("/auth/refresh")
async def refresh_token_endpoint(request: Request, response: Response):
    rt = request.cookies.get("refresh_token")
    if not rt:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(rt, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        uid = payload["sub"]
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

    pool = await get_pool()
    row = await pool.fetchrow(
        "SELECT id, email FROM users WHERE id = $1", uuid.UUID(uid)
    )
    if not row:
        raise HTTPException(status_code=401, detail="User not found")
    new_access = create_access_token(str(row["id"]), row["email"])
    response.set_cookie(
        key="access_token",
        value=new_access,
        httponly=True,
        secure=_cookie_secure(),
        samesite=_cookie_samesite(),
        max_age=ACCESS_TOKEN_TTL_MINUTES * 60,
        path="/",
    )
    return {"ok": True}


@api.post("/auth/invite", response_model=UserPublic)
async def invite_user(
    input: InviteUserInput, user: dict = Depends(get_current_user)
):
    if user["role"] != "master_admin":
        raise HTTPException(
            status_code=403, detail="Only master_admin can invite users"
        )
    role = input.role or "admin"
    if role not in ("admin", "master_admin"):
        raise HTTPException(status_code=400, detail="Invalid role")
    principal: Optional[str] = None
    if role == "admin":
        if not input.principal or input.principal not in VALID_PRINCIPALS:
            raise HTTPException(
                status_code=400,
                detail=f"principal is required for role=admin and must be one of {sorted(VALID_PRINCIPALS)}",
            )
        principal = input.principal

    email = input.email.lower()
    pool = await get_pool()
    existing = await pool.fetchval("SELECT 1 FROM users WHERE email = $1", email)
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    row = await pool.fetchrow(
        """
        INSERT INTO users (email, password_hash, name, role, principal, created_at)
        VALUES ($1, $2, $3, $4, $5, now())
        RETURNING id, email, name, role, principal, created_at
        """,
        email,
        hash_password(input.password),
        input.name,
        role,
        principal,
    )
    return {
        "_id": str(row["id"]),
        "email": row["email"],
        "name": row["name"],
        "role": row["role"],
        "principal": row["principal"],
        "created_at": row["created_at"],
    }


@api.get("/users", response_model=List[UserPublic])
async def list_users(user: dict = Depends(get_current_user)):
    if user["role"] != "master_admin":
        raise HTTPException(
            status_code=403, detail="Only master_admin can list users"
        )
    pool = await get_pool()
    rows = await pool.fetch(
        "SELECT id, email, name, role, principal, created_at FROM users ORDER BY created_at DESC"
    )
    return [
        {
            "_id": str(r["id"]),
            "email": r["email"],
            "name": r["name"],
            "role": r["role"],
            "principal": r["principal"],
            "created_at": r["created_at"],
        }
        for r in rows
    ]


# ---------- Price Record Routes ----------
PR_COLS = (
    "id, part_no, unit_price, cpq_number, cpq_date, customer, "
    "cpq_price, qty, description, notes, principal, created_by, created_by_name, created_at, updated_at"
)


@api.post("/price-records")
async def create_price_record(
    input: PriceRecordCreate, user: dict = Depends(get_current_user)
):
    principal = resolve_principal(user, input.principal)
    pool = await get_pool()
    row = await pool.fetchrow(
        f"""
        INSERT INTO price_records (
            part_no, unit_price, cpq_number, cpq_date, customer,
            cpq_price, qty, description, notes, principal, created_by, created_by_name, created_at, updated_at
        )
        VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, now(), now())
        RETURNING {PR_COLS}
        """,
        input.part_no.strip(),
        float(input.unit_price),
        input.cpq_number.strip(),
        parse_iso_date(input.cpq_date),
        input.customer.strip(),
        float(input.cpq_price),
        input.qty,
        input.description or "",
        input.notes or "",
        principal,
        uuid.UUID(user["_id"]),
        user.get("name") or user.get("email"),
    )
    return serialize_price_record(row)


@api.post("/price-records/batch")
async def create_batch(
    input: CPQBatchCreate, user: dict = Depends(get_current_user)
):
    if not input.lines:
        raise HTTPException(status_code=400, detail="No line items provided")
    principal = resolve_principal(user, input.principal)
    cpq_date = parse_iso_date(input.cpq_date)
    cpq_number = input.cpq_number.strip()
    created_by = uuid.UUID(user["_id"])
    created_by_name = user.get("name") or user.get("email")
    values = [
        (
            line.part_no.strip(),
            float(line.unit_price),
            cpq_number,
            cpq_date,
            line.customer.strip(),
            float(line.cpq_price),
            line.qty,
            line.description or "",
            line.notes or "",
            principal,
            created_by,
            created_by_name,
        )
        for line in input.lines
    ]
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.executemany(
                """
                INSERT INTO price_records (
                    part_no, unit_price, cpq_number, cpq_date, customer,
                    cpq_price, qty, description, notes, principal, created_by, created_by_name, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12, now(), now())
                """,
                values,
            )
    return {"inserted": len(values)}


@api.get("/price-records")
async def list_price_records(
    q: Optional[str] = None,
    limit: int = Query(default=100, le=1000),
    user: dict = Depends(get_current_user),
):
    pool = await get_pool()
    params: list = []
    where_clauses = []
    if q:
        params.append(f"%{q.strip()}%")
        where_clauses.append(
            f"(part_no ILIKE ${len(params)} OR cpq_number ILIKE ${len(params)} OR customer ILIKE ${len(params)})"
        )
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_clauses.append(pf_clause)
        params.extend(pf_params)
    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    params.append(limit)
    sql = f"""
        SELECT {PR_COLS} FROM price_records
        {where_sql}
        ORDER BY cpq_date DESC NULLS LAST, created_at DESC
        LIMIT ${len(params)}
    """
    rows = await pool.fetch(sql, *params)
    return [serialize_price_record(r) for r in rows]


@api.get("/price-records/parts")
async def list_parts(user: dict = Depends(get_current_user)):
    pool = await get_pool()
    pf_clause, pf_params = principal_filter_clause(user, 1)
    where_sql = f"WHERE {pf_clause}" if pf_clause else ""
    rows = await pool.fetch(
        f"""
        WITH ranked AS (
          SELECT part_no, unit_price, cpq_date, created_at,
                 ROW_NUMBER() OVER (
                   PARTITION BY part_no
                   ORDER BY cpq_date DESC NULLS LAST, created_at DESC
                 ) AS rn,
                 COUNT(*) OVER (PARTITION BY part_no) AS cnt
          FROM price_records
          {where_sql}
        )
        SELECT part_no, unit_price AS latest_unit_price,
               cpq_date AS latest_cpq_date, cnt AS record_count
        FROM ranked WHERE rn = 1
        ORDER BY part_no ASC
        """,
        *pf_params,
    )
    return [
        {
            "part_no": r["part_no"],
            "latest_unit_price": float(r["latest_unit_price"] or 0),
            "latest_cpq_date": _fmt_date(r["latest_cpq_date"]),
            "record_count": int(r["record_count"]),
        }
        for r in rows
    ]


@api.get("/price-records/by-part/{part_no}")
async def get_by_part(part_no: str, user: dict = Depends(get_current_user)):
    pool = await get_pool()
    params: list = [part_no]
    where_sql = "WHERE part_no = $1"
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_sql += f" AND {pf_clause}"
        params.extend(pf_params)
    rows = await pool.fetch(
        f"""
        SELECT {PR_COLS} FROM price_records
        {where_sql}
        ORDER BY cpq_date DESC NULLS LAST, created_at DESC
        """,
        *params,
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Part not found")
    records = [serialize_price_record(r) for r in rows]
    return {
        "part_no": part_no,
        "latest_unit_price": records[0]["unit_price"],
        "latest_cpq_date": records[0]["cpq_date"],
        "records": records,
    }


@api.get("/price-records/{record_id}")
async def get_price_record(record_id: str, user: dict = Depends(get_current_user)):
    try:
        rid = uuid.UUID(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    pool = await get_pool()
    params: list = [rid]
    where_sql = "WHERE id = $1"
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_sql += f" AND {pf_clause}"
        params.extend(pf_params)
    row = await pool.fetchrow(
        f"SELECT {PR_COLS} FROM price_records {where_sql}", *params
    )
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    return serialize_price_record(row)


@api.patch("/price-records/{record_id}")
async def update_price_record(
    record_id: str,
    input: PriceRecordUpdate,
    user: dict = Depends(get_current_user),
):
    try:
        rid = uuid.UUID(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")

    updates = {k: v for k, v in input.model_dump(exclude_unset=True).items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")

    normalized: dict = {}
    for k, v in updates.items():
        if k == "cpq_date":
            normalized[k] = parse_iso_date(v)
        elif k in ("unit_price", "cpq_price"):
            normalized[k] = float(v)
        elif k == "qty":
            normalized[k] = int(v)
        elif k in ("part_no", "cpq_number", "customer") and isinstance(v, str):
            normalized[k] = v.strip()
        else:
            normalized[k] = v

    set_clauses = []
    params: list = []
    for i, (k, v) in enumerate(normalized.items(), 1):
        set_clauses.append(f"{k} = ${i}")
        params.append(v)
    set_clauses.append(f"updated_at = now()")
    params.append(rid)
    where_sql = f"WHERE id = ${len(params)}"
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_sql += f" AND {pf_clause}"
        params.extend(pf_params)

    pool = await get_pool()
    row = await pool.fetchrow(
        f"""
        UPDATE price_records SET {', '.join(set_clauses)}
        {where_sql}
        RETURNING {PR_COLS}
        """,
        *params,
    )
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    return serialize_price_record(row)


@api.delete("/price-records/{record_id}")
async def delete_price_record(
    record_id: str, user: dict = Depends(get_current_user)
):
    try:
        rid = uuid.UUID(record_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid id")
    pool = await get_pool()
    params: list = [rid]
    where_sql = "WHERE id = $1"
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_sql += f" AND {pf_clause}"
        params.extend(pf_params)
    n = await pool.execute(f"DELETE FROM price_records {where_sql}", *params)
    # asyncpg returns "DELETE <n>"
    deleted = int(n.split()[-1]) if isinstance(n, str) else 0
    if deleted == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---------- Duplicate CPQ across customers ----------
@api.post("/price-records/duplicate")
async def duplicate_cpq(
    input: DuplicateCPQInput, user: dict = Depends(get_current_user)
):
    pool = await get_pool()
    params: list = [input.cpq_number, input.source_customer]
    where_sql = "WHERE cpq_number = $1 AND customer = $2"
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_sql += f" AND {pf_clause}"
        params.extend(pf_params)
    source_rows = await pool.fetch(
        f"""
        SELECT {PR_COLS} FROM price_records
        {where_sql}
        """,
        *params,
    )
    if not source_rows:
        raise HTTPException(
            status_code=404,
            detail=f"No records found for CPQ '{input.cpq_number}' + customer '{input.source_customer}'",
        )
    targets = [c.strip() for c in input.target_customers if c and c.strip()]
    if not targets:
        raise HTTPException(status_code=400, detail="No target customers")

    new_cpq_number = (
        input.new_cpq_number.strip() if input.new_cpq_number else input.cpq_number
    )
    new_cpq_date = (
        parse_iso_date(input.new_cpq_date)
        if input.new_cpq_date
        else source_rows[0]["cpq_date"]
    )
    created_by = uuid.UUID(user["_id"])
    created_by_name = user.get("name") or user.get("email")

    values = []
    for target in targets:
        for src in source_rows:
            values.append(
                (
                    src["part_no"],
                    float(src["unit_price"] or 0),
                    new_cpq_number,
                    new_cpq_date,
                    target,
                    float(src["cpq_price"] or 0),
                    int(src["qty"]) if src["qty"] is not None else 1,
                    src["description"] or "",
                    src["notes"] or "",
                    src["principal"],
                    created_by,
                    created_by_name,
                )
            )
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.executemany(
                """
                INSERT INTO price_records (
                    part_no, unit_price, cpq_number, cpq_date, customer,
                    cpq_price, qty, description, notes, principal, created_by, created_by_name, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12, now(), now())
                """,
                values,
            )
    return {
        "inserted": len(values),
        "new_cpq_number": new_cpq_number,
        "target_customers": targets,
    }


# ---------- Excel Import ----------
@api.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...), user: dict = Depends(get_current_user)
):
    if not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Upload .xlsx, .xls or .csv")
    content = await file.read()
    try:
        if file.filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content))
        else:
            df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")
    df = df.fillna("")
    columns = [str(c) for c in df.columns.tolist()]
    sample = df.head(20).astype(str).to_dict(orient="records")
    all_rows = df.astype(str).to_dict(orient="records")
    return {
        "columns": columns,
        "sample_rows": sample,
        "all_rows": all_rows,
        "row_count": len(all_rows),
    }


@api.post("/import/commit")
async def import_commit(
    payload: ImportCommit, user: dict = Depends(get_current_user)
):
    principal = resolve_principal(user, payload.principal)
    errors: List[dict] = []
    values = []
    created_by = uuid.UUID(user["_id"])
    created_by_name = user.get("name") or user.get("email")
    for idx, row in enumerate(payload.rows):
        try:
            if not (row.cpq_price or "").strip():
                raise ValueError("CPQ Price is required")
            # unit_price (list price) may be unknown at import time — e.g. a
            # Schneider PDF quote never contains it — so default to 0 and let
            # the user fill it in later via the part's Edit dialog.
            unit_price_raw = (row.unit_price or "").strip()
            unit_price = float(unit_price_raw) if unit_price_raw else 0.0
            values.append(
                (
                    row.part_no.strip(),
                    unit_price,
                    row.cpq_number.strip(),
                    parse_iso_date(row.cpq_date),
                    row.customer.strip(),
                    float(row.cpq_price),
                    parse_qty(row.qty),
                    row.description or "",
                    row.notes or "",
                    principal,
                    created_by,
                    created_by_name,
                )
            )
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})
    if not values:
        preview = "; ".join(f"row {e['row'] + 1}: {e['error']}" for e in errors[:5])
        more = f" (+{len(errors) - 5} more)" if len(errors) > 5 else ""
        raise HTTPException(
            status_code=400,
            detail=f"No valid rows imported ({len(errors)} error(s)). {preview}{more}",
        )
    pool = await get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            await conn.executemany(
                """
                INSERT INTO price_records (
                    part_no, unit_price, cpq_number, cpq_date, customer,
                    cpq_price, qty, description, notes, principal, created_by, created_by_name, created_at, updated_at
                )
                VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12, now(), now())
                """,
                values,
            )
    return {"inserted": len(values), "errors": errors}


@api.post("/import/pdf")
async def import_pdf(
    file: UploadFile = File(...), user: dict = Depends(get_current_user)
):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Upload a .pdf file")
    content = await file.read()
    try:
        parsed = parse_schneider_pdf(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {e}")
    rows = parsed["rows"]
    if not rows:
        raise HTTPException(status_code=400, detail="No line items found in PDF")
    columns = [
        "part_no",
        "description",
        "qty",
        "unit_price",
        "cpq_number",
        "cpq_date",
        "customer",
        "cpq_price",
    ]
    return {
        "columns": columns,
        "sample_rows": rows[:20],
        "all_rows": rows,
        "row_count": len(rows),
    }


@api.get("/stats")
async def stats(user: dict = Depends(get_current_user)):
    pool = await get_pool()
    pf_clause, pf_params = principal_filter_clause(user, 1)
    where_sql = f"WHERE {pf_clause}" if pf_clause else ""
    total_records = await pool.fetchval(
        f"SELECT COUNT(*) FROM price_records {where_sql}", *pf_params
    )
    distinct_parts = await pool.fetchval(
        f"SELECT COUNT(DISTINCT part_no) FROM price_records {where_sql}", *pf_params
    )
    distinct_customers = await pool.fetchval(
        f"SELECT COUNT(DISTINCT customer) FROM price_records {where_sql}", *pf_params
    )
    distinct_cpq = await pool.fetchval(
        f"SELECT COUNT(DISTINCT cpq_number) FROM price_records {where_sql}", *pf_params
    )
    return {
        "total_records": int(total_records or 0),
        "distinct_parts": int(distinct_parts or 0),
        "distinct_customers": int(distinct_customers or 0),
        "distinct_cpq": int(distinct_cpq or 0),
    }


# ---------- Excel Export ----------
_XLSX_HEADERS = [
    "Part No",
    "Description",
    "Qty",
    "CPQ #",
    "CPQ Date",
    "Customer",
    "Unit Price (RM)",
    "CPQ Price (RM)",
    "Discount %",
    "Notes",
    "Created By",
    "Created At",
]
_XLSX_COL_WIDTHS = [16, 30, 8, 18, 12, 22, 16, 16, 12, 30, 18, 18]


def _xlsx_write_header(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    for col, h in enumerate(_XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _xlsx_format_created(created: Any) -> str:
    if isinstance(created, datetime):
        return created.strftime("%Y-%m-%d %H:%M")
    return str(created or "")


def _xlsx_write_row(ws, row_idx: int, r: dict) -> None:
    values = [
        r.get("part_no", ""),
        r.get("description", ""),
        int(r.get("qty", 1) or 1),
        r.get("cpq_number", ""),
        r.get("cpq_date", ""),
        r.get("customer", ""),
        float(r.get("unit_price", 0) or 0),
        float(r.get("cpq_price", 0) or 0),
        float(r.get("discount_pct", 0) or 0),
        r.get("notes", ""),
        r.get("created_by_name", ""),
        _xlsx_format_created(r.get("created_at")),
    ]
    for col, val in enumerate(values, 1):
        ws.cell(row=row_idx, column=col, value=val)


def _xlsx_apply_styles(ws) -> None:
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = '"RM " #,##0.00'
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.number_format = "0.00%"
    for i, w in enumerate(_XLSX_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_xlsx(rows: List[dict], title: str = "Price Records") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _xlsx_write_header(ws)
    for r_idx, r in enumerate(rows, 2):
        _xlsx_write_row(ws, r_idx, r)
    _xlsx_apply_styles(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@api.get("/export/xlsx")
async def export_xlsx(
    q: Optional[str] = None,
    part_no: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    pool = await get_pool()
    params: list = []
    where_clauses = []
    if part_no:
        params.append(part_no)
        where_clauses.append(f"part_no = ${len(params)}")
    elif q:
        params.append(f"%{q.strip()}%")
        where_clauses.append(
            f"(part_no ILIKE ${len(params)} OR cpq_number ILIKE ${len(params)} OR customer ILIKE ${len(params)})"
        )
    pf_clause, pf_params = principal_filter_clause(user, len(params) + 1)
    if pf_clause:
        where_clauses.append(pf_clause)
        params.extend(pf_params)
    where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    rows = await pool.fetch(
        f"""
        SELECT {PR_COLS} FROM price_records
        {where_sql}
        ORDER BY cpq_date DESC NULLS LAST, created_at DESC
        LIMIT 10000
        """,
        *params,
    )

    serialized = [serialize_price_record(r) for r in rows]
    title = f"Part {part_no}" if part_no else "Price Records"
    xlsx_bytes = _build_xlsx(serialized, title=title)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    filename = (
        f"pricepoint-part-{part_no}-{stamp}.xlsx"
        if part_no
        else f"pricepoint-price-records-{stamp}.xlsx"
    )
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------- Mount ----------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)
