"""Backend tests for new features: Excel export + CPQ duplicate."""
import os
import uuid
import pytest
import requests


BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip()
                break
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = os.environ.get("TEST_ADMIN_EMAIL", "admin@farg.com")
ADMIN_PASSWORD = os.environ.get("TEST_ADMIN_PASSWORD", "admin123")


@pytest.fixture(scope="module")
def admin_session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, r.text
    return s


@pytest.fixture(scope="module")
def seeded_cpq(admin_session):
    """Seed a CPQ with source customer 'Petronas' having 3 line items."""
    cpq_no = f"TEST-DUP-CPQ-{uuid.uuid4().hex[:6]}"
    payload = {
        "cpq_number": cpq_no,
        "cpq_date": "2025-06-15",
        "lines": [
            {"part_no": f"TEST-DUP-P1-{uuid.uuid4().hex[:4]}", "unit_price": 1000, "customer": "Petronas", "cpq_price": 850, "notes": "n1"},
            {"part_no": f"TEST-DUP-P2-{uuid.uuid4().hex[:4]}", "unit_price": 2000, "customer": "Petronas", "cpq_price": 1700, "notes": "n2"},
            {"part_no": f"TEST-DUP-P3-{uuid.uuid4().hex[:4]}", "unit_price": 500, "customer": "Petronas", "cpq_price": 450, "notes": ""},
        ],
    }
    r = admin_session.post(f"{API}/price-records/batch", json=payload)
    assert r.status_code == 200
    return {"cpq_number": cpq_no, "source_customer": "Petronas", "line_count": 3}


# -------------------- Excel Export --------------------
class TestExportXlsx:
    def test_export_unauth_returns_401(self):
        r = requests.get(f"{API}/export/xlsx")
        assert r.status_code == 401

    def test_export_all_records(self, admin_session):
        r = admin_session.get(f"{API}/export/xlsx")
        assert r.status_code == 200
        # Content-Type
        assert r.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Zip magic bytes
        assert r.content[:2] == b"PK", f"bad magic bytes: {r.content[:4]!r}"
        # Content-Disposition
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower()
        assert ".xlsx" in cd.lower()
        # Not tiny — should have some content
        assert len(r.content) > 500

    def test_export_with_q_filter(self, admin_session, seeded_cpq):
        # Filter by CPQ number should return only matching rows
        r = admin_session.get(f"{API}/export/xlsx", params={"q": seeded_cpq["cpq_number"]})
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        assert r.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # We can open the xlsx and verify it has our rows
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # header + N data rows
        assert len(rows) >= 1 + seeded_cpq["line_count"]
        # All non-header rows for this filter should have the matching CPQ #
        data_rows = rows[1:]
        cpq_col_values = [row[1] for row in data_rows]  # CPQ # is column 2
        assert all(v == seeded_cpq["cpq_number"] for v in cpq_col_values), cpq_col_values

    def test_export_with_part_no(self, admin_session, seeded_cpq, admin_session_extra_part):
        part = admin_session_extra_part
        r = admin_session.get(f"{API}/export/xlsx", params={"part_no": part})
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # header + at least 1 record for this part
        assert len(rows) >= 2
        # column 1 is Part No
        part_col_values = [row[0] for row in rows[1:]]
        assert all(v == part for v in part_col_values), part_col_values

    def test_export_filename_scoped_to_part(self, admin_session, admin_session_extra_part):
        r = admin_session.get(f"{API}/export/xlsx", params={"part_no": admin_session_extra_part})
        cd = r.headers.get("content-disposition", "")
        assert admin_session_extra_part in cd


@pytest.fixture(scope="module")
def admin_session_extra_part(admin_session):
    """Create a specific part for the part_no filter test."""
    part = f"TEST-EXPORTPART-{uuid.uuid4().hex[:6]}"
    r = admin_session.post(
        f"{API}/price-records",
        json={
            "part_no": part,
            "unit_price": 1234.0,
            "cpq_number": f"CPQ-EXP-{uuid.uuid4().hex[:4]}",
            "cpq_date": "2025-07-01",
            "customer": "ExpCust",
            "cpq_price": 999.0,
        },
    )
    assert r.status_code == 200
    return part


# -------------------- Duplicate CPQ --------------------
class TestDuplicateCPQ:
    def test_duplicate_success(self, admin_session, seeded_cpq):
        payload = {
            "cpq_number": seeded_cpq["cpq_number"],
            "source_customer": seeded_cpq["source_customer"],
            "target_customers": ["Shell", "ExxonMobil"],
        }
        r = admin_session.post(f"{API}/price-records/duplicate", json=payload)
        assert r.status_code == 200, r.text
        data = r.json()
        # 3 line items x 2 targets = 6 new rows
        assert data["inserted"] == seeded_cpq["line_count"] * 2
        assert set(data["target_customers"]) == {"Shell", "ExxonMobil"}
        assert data["new_cpq_number"] == seeded_cpq["cpq_number"]

        # Verify via search - now should have records for Shell + ExxonMobil under same CPQ
        r2 = admin_session.get(f"{API}/price-records", params={"q": seeded_cpq["cpq_number"]})
        assert r2.status_code == 200
        rows = r2.json()
        customers = {row["customer"] for row in rows}
        assert {"Petronas", "Shell", "ExxonMobil"}.issubset(customers)

        # Original Petronas records should still exist unchanged (line_count of them)
        petronas_rows = [row for row in rows if row["customer"] == "Petronas"]
        assert len(petronas_rows) == seeded_cpq["line_count"]

    def test_duplicate_with_new_cpq_number_and_date(self, admin_session, seeded_cpq):
        new_cpq_no = f"TEST-DUP-NEW-{uuid.uuid4().hex[:6]}"
        payload = {
            "cpq_number": seeded_cpq["cpq_number"],
            "source_customer": seeded_cpq["source_customer"],
            "target_customers": ["Chevron"],
            "new_cpq_number": new_cpq_no,
            "new_cpq_date": "2026-01-10",
        }
        r = admin_session.post(f"{API}/price-records/duplicate", json=payload)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["new_cpq_number"] == new_cpq_no
        assert data["inserted"] == seeded_cpq["line_count"]

        # Verify new CPQ number is applied
        r2 = admin_session.get(f"{API}/price-records", params={"q": new_cpq_no})
        assert r2.status_code == 200
        rows = r2.json()
        assert len(rows) == seeded_cpq["line_count"]
        for row in rows:
            assert row["customer"] == "Chevron"
            assert row["cpq_number"] == new_cpq_no
            assert row["cpq_date"] == "2026-01-10"

    def test_duplicate_source_not_found(self, admin_session):
        payload = {
            "cpq_number": "NONEXISTENT-CPQ-XYZ",
            "source_customer": "NoSuchCustomer",
            "target_customers": ["Shell"],
        }
        r = admin_session.post(f"{API}/price-records/duplicate", json=payload)
        assert r.status_code == 404

    def test_duplicate_empty_target_customers(self, admin_session, seeded_cpq):
        payload = {
            "cpq_number": seeded_cpq["cpq_number"],
            "source_customer": seeded_cpq["source_customer"],
            "target_customers": [],
        }
        r = admin_session.post(f"{API}/price-records/duplicate", json=payload)
        # Pydantic min_length=1 -> 422
        assert r.status_code in (400, 422)

    def test_duplicate_blank_target_customers(self, admin_session, seeded_cpq):
        payload = {
            "cpq_number": seeded_cpq["cpq_number"],
            "source_customer": seeded_cpq["source_customer"],
            "target_customers": ["   ", ""],
        }
        r = admin_session.post(f"{API}/price-records/duplicate", json=payload)
        # After trimming, no valid targets -> 400
        assert r.status_code == 400

    def test_duplicate_unauth(self):
        r = requests.post(
            f"{API}/price-records/duplicate",
            json={
                "cpq_number": "X",
                "source_customer": "Y",
                "target_customers": ["Z"],
            },
        )
        assert r.status_code == 401
